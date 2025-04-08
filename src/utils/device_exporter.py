import csv
import logging
import os
from datetime import datetime

class DeviceExporter:
    """Utility class for exporting devices to various formats."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.logger = logging.getLogger(__name__)
    
    def export_to_csv(self, devices, filepath=None):
        """
        Export a list of devices to CSV format.
        
        Args:
            devices: List of Device objects to export
            filepath: Optional filepath to save the CSV. If None, a default path is generated.
            
        Returns:
            str: Path to the saved CSV file
        """
        if not devices:
            self.logger.warning("No devices to export")
            return None
            
        # Generate default filename if not provided
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"device_export_{timestamp}.csv"
        
        # Collect all possible property names from all devices
        property_names = set()
        for device in devices:
            if hasattr(device, 'properties') and isinstance(device.properties, dict):
                property_names.update(device.properties.keys())
            else:
                self.logger.warning(f"Device {getattr(device, 'name', 'unknown')} has invalid properties attribute")
        
        # Sort property names alphabetically for consistent output
        property_names = sorted(list(property_names))
        
        # Define the CSV header row
        header = ["id", "name", "device_type", "x_position", "y_position"] + property_names
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header row
                writer.writerow(header)
                
                # Write data rows
                for device in devices:
                    try:
                        # Get position with error handling
                        try:
                            pos = device.scenePos()
                            x_pos = pos.x()
                            y_pos = pos.y()
                        except (AttributeError, TypeError) as e:
                            self.logger.warning(f"Error getting position for device {getattr(device, 'name', 'unknown')}: {str(e)}")
                            x_pos = 0
                            y_pos = 0
                        
                        # Build row starting with fixed fields
                        row = [
                            getattr(device, 'id', f"unknown-{id(device)}"),
                            getattr(device, 'name', "Unknown Device"),
                            getattr(device, 'device_type', "GENERIC"),
                            x_pos,
                            y_pos
                        ]
                        
                        # Add properties in same order as header with validation
                        if hasattr(device, 'properties') and isinstance(device.properties, dict):
                            for prop in property_names:
                                # Get property value with default
                                value = device.properties.get(prop, "")
                                # Convert value to string if not None
                                str_value = str(value) if value is not None else ""
                                row.append(str_value)
                        else:
                            # If properties is invalid, fill with empty strings
                            row.extend([""] * len(property_names))
                        
                        writer.writerow(row)
                    except Exception as row_error:
                        self.logger.error(f"Error processing device {getattr(device, 'name', 'unknown')}: {str(row_error)}")
                        continue
                    
            self.logger.info(f"Exported {len(devices)} devices to {filepath}")
            return filepath
            
        except (IOError, PermissionError) as e:
            self.logger.error(f"File error when exporting to CSV: {str(e)}")
            return None
        except Exception as e:
            self.logger.error(f"Error exporting devices to CSV: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None
    
    def export_to_excel(self, devices, filepath=None):
        """
        Export devices to Excel format.
        
        Args:
            devices: List of Device objects to export
            filepath: Optional filepath to save the Excel file. If None, a default path is generated.
            
        Returns:
            str: Path to the saved Excel file
        """
        # Check if openpyxl is available
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.logger.error("openpyxl library not found. Install it with 'pip install openpyxl'")
            return None
            
        if not devices:
            self.logger.warning("No devices to export")
            return None
            
        # Generate default filename if not provided
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"device_export_{timestamp}.xlsx"
        
        # Collect all possible property names from all devices
        property_names = set()
        for device in devices:
            if hasattr(device, 'properties') and isinstance(device.properties, dict):
                property_names.update(device.properties.keys())
            else:
                self.logger.warning(f"Device {getattr(device, 'name', 'unknown')} has invalid properties attribute")
        
        # Sort property names alphabetically for consistent output
        property_names = sorted(list(property_names))
        
        # Define the header row
        header = ["ID", "Device Name", "Device Type", "X Position", "Y Position"] + [
            prop.replace('_', ' ').title() for prop in property_names
        ]
        
        try:
            # Create a new workbook and select the active sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Devices"
            
            # Add title and export date
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = "Network Device Export"
            title_cell.font = Font(size=14, bold=True)
            
            ws.merge_cells('A2:E2')
            date_cell = ws['A2']
            date_cell.value = f"Exported on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date_cell.font = Font(italic=True)
            
            # Add header row with styling
            header_row = 4
            for col_num, column_title in enumerate(header, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_num, device in enumerate(devices, header_row + 1):
                try:
                    # Get position with error handling
                    try:
                        pos = device.scenePos()
                        x_pos = pos.x()
                        y_pos = pos.y()
                    except (AttributeError, TypeError) as e:
                        self.logger.warning(f"Error getting position for device {getattr(device, 'name', 'unknown')}: {str(e)}")
                        x_pos = 0
                        y_pos = 0
                    
                    # Add fixed fields with error handling
                    ws.cell(row=row_num, column=1).value = getattr(device, 'id', f"unknown-{row_num}")
                    ws.cell(row=row_num, column=2).value = getattr(device, 'name', f"Device-{row_num}")
                    ws.cell(row=row_num, column=3).value = getattr(device, 'device_type', "GENERIC")
                    ws.cell(row=row_num, column=4).value = x_pos
                    ws.cell(row=row_num, column=5).value = y_pos
                    
                    # Add properties with validation
                    if hasattr(device, 'properties') and isinstance(device.properties, dict):
                        for col_num, prop in enumerate(property_names, 6):
                            # Get property value with default
                            value = device.properties.get(prop, "")
                            # Convert value to string if not None
                            str_value = str(value) if value is not None else ""
                            ws.cell(row=row_num, column=col_num).value = str_value
                    else:
                        # If properties is invalid, fill with empty strings
                        for col_num in range(6, 6 + len(property_names)):
                            ws.cell(row=row_num, column=col_num).value = ""
                except Exception as row_error:
                    self.logger.error(f"Error processing device at row {row_num}: {str(row_error)}")
                    continue
            
            # Adjust column widths
            for col in ws.columns:
                try:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except (TypeError, ValueError):
                                pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
                except Exception as col_error:
                    self.logger.warning(f"Error adjusting column width: {str(col_error)}")
            
            # Save the workbook
            try:
                wb.save(filepath)
                self.logger.info(f"Exported {len(devices)} devices to Excel: {filepath}")
                return filepath
            except PermissionError:
                self.logger.error(f"Permission denied when saving Excel file: {filepath}")
                return None
            
        except Exception as e:
            self.logger.error(f"Error exporting devices to Excel: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None 