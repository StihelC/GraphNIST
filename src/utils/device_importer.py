import csv
import logging
import os
import random
from PyQt5.QtCore import QPointF

class DeviceImporter:
    """Utility class for importing devices from various file formats."""
    
    def __init__(self, device_factory=None):
        """
        Initialize the importer.
        
        Args:
            device_factory: Function that creates a device given a name, type, and properties
        """
        self.logger = logging.getLogger(__name__)
        self.device_factory = device_factory
    
    def import_from_csv(self, filepath, device_factory=None, auto_position=True):
        """
        Import devices from a CSV file.
        
        Args:
            filepath: Path to the CSV file
            device_factory: Function to create devices (overrides constructor factory)
            auto_position: If True, positions are read from file. If False, devices are arranged in a grid.
            
        Returns:
            list: List of imported devices
        """
        if device_factory is None:
            device_factory = self.device_factory
            
        if device_factory is None:
            self.logger.error("No device factory provided")
            return []
            
        # Check if file exists
        if not os.path.exists(filepath):
            self.logger.error(f"File not found: {filepath}")
            return []
            
        try:
            devices = []
            
            with open(filepath, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                
                # Parse header row
                try:
                    header = next(reader)
                except StopIteration:
                    self.logger.error("CSV file is empty")
                    return []
                
                # Extract indices for fixed fields
                try:
                    id_idx = header.index('id')
                except ValueError:
                    id_idx = None
                    self.logger.warning("No 'id' column found in CSV, will generate new IDs")
                
                try:
                    name_idx = header.index('name')
                except ValueError:
                    self.logger.error("CSV file must have a 'name' column")
                    return []
                    
                try:
                    type_idx = header.index('device_type')
                except ValueError:
                    self.logger.error("CSV file must have a 'device_type' column")
                    return []
                
                # Get position indices if they exist
                try:
                    x_pos_idx = header.index('x_position')
                    y_pos_idx = header.index('y_position')
                    has_position = True
                except ValueError:
                    self.logger.info("No position columns found in CSV, using grid layout")
                    has_position = False
                
                # Map remaining columns to properties
                property_indices = {}
                for i, column in enumerate(header):
                    if i not in [id_idx, name_idx, type_idx]:
                        if has_position and i in [x_pos_idx, y_pos_idx]:
                            continue
                        property_indices[column] = i
                
                self.logger.debug(f"Found {len(property_indices)} property columns in CSV")
                
                # Grid layout parameters if auto-positioning
                if not auto_position or not has_position:
                    grid_size = 150  # Space between devices
                    grid_cols = 5    # Number of columns in the grid
                    start_x = 100     # Starting X position
                    start_y = 100     # Starting Y position
                
                # Process each row
                row_count = 0
                for row in reader:
                    row_count += 1
                    
                    # Skip empty rows
                    if not row or len(row) < 2:
                        continue
                    
                    # Basic validation
                    if len(row) < len(header):
                        self.logger.warning(f"Row {row_count} has fewer columns than header, skipping")
                        continue
                    
                    try:
                        # Extract basic device info
                        name = row[name_idx].strip()
                        if not name:
                            self.logger.warning(f"Row {row_count} has no name, skipping")
                            continue
                            
                        device_type = row[type_idx].strip()
                        if not device_type:
                            self.logger.warning(f"Row {row_count} has no device type, using GENERIC")
                            device_type = "GENERIC"
                        
                        # Extract position
                        if auto_position and has_position:
                            try:
                                x_pos = float(row[x_pos_idx])
                                y_pos = float(row[y_pos_idx])
                            except (ValueError, IndexError):
                                # If position is invalid, calculate grid position
                                self.logger.warning(f"Invalid position for row {row_count}, using grid position")
                                col = (row_count - 1) % grid_cols
                                row_num = (row_count - 1) // grid_cols
                                x_pos = start_x + col * grid_size
                                y_pos = start_y + row_num * grid_size
                        else:
                            # Calculate grid position
                            col = (row_count - 1) % grid_cols
                            row_num = (row_count - 1) // grid_cols
                            x_pos = start_x + col * grid_size
                            y_pos = start_y + row_num * grid_size
                        
                        # Extract properties
                        properties = {}
                        for prop_name, prop_idx in property_indices.items():
                            try:
                                value = row[prop_idx]
                                if value:  # Only add non-empty properties
                                    properties[prop_name] = value
                            except IndexError:
                                pass
                        
                        # Create the device
                        try:
                            device = device_factory(name, device_type, properties)
                            
                            if device:
                                # Set position
                                device.setPos(QPointF(x_pos, y_pos))
                                devices.append(device)
                                self.logger.debug(f"Created device '{name}' of type '{device_type}' at ({x_pos}, {y_pos})")
                            else:
                                self.logger.warning(f"Failed to create device for row {row_count}")
                        except Exception as e:
                            self.logger.error(f"Error creating device for row {row_count}: {str(e)}")
                            continue
                            
                    except Exception as row_e:
                        self.logger.error(f"Error processing row {row_count}: {str(row_e)}")
                        continue
                    
            self.logger.info(f"Imported {len(devices)} devices from {filepath}")
            return devices
                        
        except Exception as e:
            self.logger.error(f"Error importing devices from CSV: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return []
    
    def import_from_excel(self, filepath, device_factory=None, auto_position=True, sheet_name=None):
        """
        Import devices from an Excel file.
        
        Args:
            filepath: Path to the Excel file
            device_factory: Function to create devices (overrides constructor factory)
            auto_position: If True, positions are read from file. If False, devices are arranged in a grid.
            sheet_name: Name of the sheet to read from (if None, the first sheet is used)
            
        Returns:
            list: List of imported devices
        """
        # Check if openpyxl is available
        try:
            import openpyxl
        except ImportError:
            self.logger.error("openpyxl library not found. Install it with 'pip install openpyxl'")
            return []
            
        if device_factory is None:
            device_factory = self.device_factory
            
        if device_factory is None:
            self.logger.error("No device factory provided")
            return []
            
        # Check if file exists
        if not os.path.exists(filepath):
            self.logger.error(f"File not found: {filepath}")
            return []
            
        try:
            devices = []
            
            # Load the workbook
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            
            # Select the worksheet
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self.logger.info(f"Using sheet '{sheet_name}'")
            else:
                ws = wb.active
                self.logger.info(f"Using active sheet '{ws.title}'")
            
            # Get the header row
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
                # Look for rows with "name" and "device_type" as these are required
                if any(cell.value == "name" or cell.value == "Device Name" for cell in row) and \
                   any(cell.value == "device_type" or cell.value == "Device Type" for cell in row):
                    header_row = [cell.value for cell in row]
                    self.logger.debug(f"Found header row at row {row_idx + 1}")
                    break
                    
            if not header_row:
                self.logger.error("Could not find a valid header row with 'name' and 'device_type' columns")
                return []
            
            # Normalize header names
            normalized_header = []
            for column in header_row:
                if not column:
                    normalized_header.append("")
                    continue
                    
                # Convert to lowercase and replace spaces with underscores
                if isinstance(column, str):
                    column = column.lower().replace(' ', '_')
                normalized_header.append(column)
            
            # Extract indices for fixed fields
            try:
                id_idx = normalized_header.index('id')
            except ValueError:
                try:
                    id_idx = normalized_header.index('device_id')
                except ValueError:
                    id_idx = None
                    self.logger.warning("No 'id' column found in Excel, will generate new IDs")
            
            try:
                name_idx = normalized_header.index('name')
            except ValueError:
                try:
                    name_idx = normalized_header.index('device_name')
                except ValueError:
                    self.logger.error("Excel file must have a 'name' or 'device_name' column")
                    return []
                    
            try:
                type_idx = normalized_header.index('device_type')
            except ValueError:
                try:
                    type_idx = normalized_header.index('type')
                except ValueError:
                    self.logger.error("Excel file must have a 'device_type' or 'type' column")
                    return []
            
            # Get position indices if they exist
            try:
                x_pos_idx = normalized_header.index('x_position')
                y_pos_idx = normalized_header.index('y_position')
                has_position = True
            except ValueError:
                try:
                    x_pos_idx = normalized_header.index('x')
                    y_pos_idx = normalized_header.index('y')
                    has_position = True
                except ValueError:
                    self.logger.info("No position columns found in Excel, using grid layout")
                    has_position = False
            
            # Map remaining columns to properties
            property_indices = {}
            for i, column in enumerate(normalized_header):
                if i not in [id_idx, name_idx, type_idx]:
                    if column and (not has_position or (i != x_pos_idx and i != y_pos_idx)):
                        property_indices[column] = i
            
            self.logger.debug(f"Found {len(property_indices)} property columns in Excel")
            
            # Grid layout parameters if auto-positioning
            if not auto_position or not has_position:
                grid_size = 150  # Space between devices
                grid_cols = 5    # Number of columns in the grid
                start_x = 100     # Starting X position
                start_y = 100     # Starting Y position
            
            # Find the header row index to start reading data from
            header_index = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
                if any(cell.value == "name" or cell.value == "Device Name" for cell in row):
                    header_index = i + 1
                    break
            
            if header_index is None:
                self.logger.error("Could not find header row")
                return []
                
            # Process each data row
            row_count = 0
            for i, row in enumerate(ws.iter_rows(min_row=header_index+1), 1):
                row_values = [cell.value for cell in row]
                
                # Skip empty rows
                if not any(row_values):
                    continue
                
                row_count += 1
                
                try:
                    # Extract basic device info
                    try:
                        name = row_values[name_idx]
                        if not name:  # Skip rows without names
                            self.logger.warning(f"Row {i} has no name, skipping")
                            continue
                            
                        device_type = row_values[type_idx]
                        if not device_type:  # Use default if empty
                            self.logger.warning(f"Row {i} has no device type, using GENERIC")
                            device_type = "GENERIC"
                    except IndexError:
                        self.logger.warning(f"Row {i} is invalid, skipping")
                        continue
                    
                    # Extract position
                    if auto_position and has_position:
                        try:
                            x_pos = float(row_values[x_pos_idx] or 0)
                            y_pos = float(row_values[y_pos_idx] or 0)
                        except (ValueError, IndexError, TypeError):
                            # If position is invalid, calculate grid position
                            self.logger.warning(f"Invalid position for row {i}, using grid position")
                            col = (row_count - 1) % grid_cols
                            row_num = (row_count - 1) // grid_cols
                            x_pos = start_x + col * grid_size
                            y_pos = start_y + row_num * grid_size
                    else:
                        # Calculate grid position
                        col = (row_count - 1) % grid_cols
                        row_num = (row_count - 1) // grid_cols
                        x_pos = start_x + col * grid_size
                        y_pos = start_y + row_num * grid_size
                    
                    # Extract properties
                    properties = {}
                    for prop_name, prop_idx in property_indices.items():
                        try:
                            value = row_values[prop_idx]
                            if value is not None:  # Only add non-empty properties
                                properties[prop_name] = str(value)
                        except IndexError:
                            pass
                    
                    # Create the device
                    try:
                        device = device_factory(name, device_type, properties)
                        
                        if device:
                            # Set position
                            device.setPos(QPointF(x_pos, y_pos))
                            devices.append(device)
                            self.logger.debug(f"Created device '{name}' of type '{device_type}' at ({x_pos}, {y_pos})")
                        else:
                            self.logger.warning(f"Failed to create device for row {i}")
                    except Exception as e:
                        self.logger.error(f"Error creating device for row {i}: {str(e)}")
                        continue
                        
                except Exception as row_e:
                    self.logger.error(f"Error processing row {i}: {str(row_e)}")
                    continue
            
            self.logger.info(f"Imported {len(devices)} devices from Excel: {filepath}")
            return devices
            
        except Exception as e:
            self.logger.error(f"Error importing devices from Excel: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return []
    
    def import_from_json(self, filepath, device_factory=None):
        """
        Import devices from a JSON file.
        
        Args:
            filepath: Path to the JSON file
            device_factory: Function to create devices (overrides constructor factory)
            
        Returns:
            list: List of imported devices
        """
        import json
        
        if device_factory is None:
            device_factory = self.device_factory
            
        if device_factory is None:
            self.logger.error("No device factory provided")
            return []
            
        # Check if file exists
        if not os.path.exists(filepath):
            self.logger.error(f"File not found: {filepath}")
            return []
            
        try:
            devices = []
            
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            # Validate data structure
            if not isinstance(data, list):
                if isinstance(data, dict) and 'devices' in data:
                    device_data = data['devices']
                else:
                    self.logger.error("JSON file must contain a list of devices or a 'devices' key")
                    return []
            else:
                device_data = data
                
            # Process each device
            for i, device_info in enumerate(device_data):
                if not isinstance(device_info, dict):
                    self.logger.warning(f"Device {i} is not a valid object, skipping")
                    continue
                    
                # Extract required fields
                name = device_info.get('name')
                device_type = device_info.get('device_type')
                
                if not name or not device_type:
                    self.logger.warning(f"Device {i} missing required fields, skipping")
                    continue
                
                # Extract position
                position = device_info.get('position', {})
                if isinstance(position, dict) and 'x' in position and 'y' in position:
                    x_pos = float(position['x'])
                    y_pos = float(position['y'])
                else:
                    # Assign a random position if not specified
                    x_pos = random.randint(100, 800)
                    y_pos = random.randint(100, 600)
                
                # Extract properties
                properties = device_info.get('properties', {})
                
                # Create the device
                device = device_factory(name, device_type, properties)
                
                if device:
                    # Set position
                    device.setPos(QPointF(x_pos, y_pos))
                    devices.append(device)
                    
            self.logger.info(f"Imported {len(devices)} devices from JSON: {filepath}")
            return devices
            
        except Exception as e:
            self.logger.error(f"Error importing devices from JSON: {str(e)}")
            return [] 